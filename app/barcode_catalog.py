"""Load and query the Barcode Master List Excel file."""

from __future__ import annotations

import json
import shutil
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.paths import get_data_dir
from app.pdf_parser import normalize_part


def _data_dir() -> Path:
    return get_data_dir()


def _db_path() -> Path:
    return _data_dir() / "scanner.db"


def _config_path() -> Path:
    return _data_dir() / "config.json"


def _master_path() -> Path:
    return _data_dir() / "BarcodeMasterList.xlsx"


BARCODE_HEADERS = {"barcode"}
PART_HEADERS = {"item part nubmer", "item part number", "part no", "part_no", "deks part #"}
DESC_HEADERS = {"description"}
BOX_QTY_HEADERS = {"boxqty", "box qty", "box_qty", "carton/qty", "single /qty", "single/qty"}
PALLET_QTY_HEADERS = {"palletqty", "pallet qty", "pallet_qty"}

CATALOG_SCHEMA_VERSION = "2"


def _load_config() -> dict[str, Any]:
    config_path = _config_path()
    if not config_path.exists():
        return {}
    try:
        data = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def _save_config(config: dict[str, Any]) -> None:
    _data_dir().mkdir(parents=True, exist_ok=True)
    _config_path().write_text(json.dumps(config, indent=2), encoding="utf-8")


def get_master_path() -> Path:
    config = _load_config()
    configured = (config.get("barcode_master_path") or "").strip()
    if configured:
        path = Path(configured)
        if path.exists():
            return path
    return _master_path()


def set_default_master_path(path: Path | str) -> Path:
    """Set the default Barcode Master List path and copy it into the app data folder."""
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Barcode master list not found: {source}")
    if source.suffix.lower() != ".xlsx":
        raise ValueError("Barcode master list must be an Excel .xlsx file.")

    dest = _master_path()
    _data_dir().mkdir(parents=True, exist_ok=True)
    if source.resolve() != dest.resolve():
        shutil.copy2(source, dest)

    config = _load_config()
    config["barcode_master_path"] = str(dest)
    _save_config(config)
    return dest


def _save_config_path() -> None:
    config = _load_config()
    config["barcode_master_path"] = str(get_master_path())
    _save_config(config)


def _connect() -> sqlite3.Connection:
    _data_dir().mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(_db_path(), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def _ensure_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS barcode_master (
            barcode TEXT PRIMARY KEY,
            part_no TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            box_qty INTEGER,
            pallet_qty INTEGER
        );
        CREATE INDEX IF NOT EXISTS idx_barcode_master_part ON barcode_master(part_no);

        CREATE TABLE IF NOT EXISTS app_metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )
    columns = {
        row[1] for row in conn.execute("PRAGMA table_info(barcode_master)").fetchall()
    }
    if "box_qty" not in columns:
        conn.execute("ALTER TABLE barcode_master ADD COLUMN box_qty INTEGER")
        conn.execute("DELETE FROM app_metadata WHERE key = 'barcode_master_count'")
    if "pallet_qty" not in columns:
        conn.execute("ALTER TABLE barcode_master ADD COLUMN pallet_qty INTEGER")
        conn.execute("DELETE FROM app_metadata WHERE key = 'barcode_master_count'")


def _column_map(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header is None:
            continue
        name = str(header).strip().lower()
        if name in BARCODE_HEADERS:
            mapping["barcode"] = index
        elif name in PART_HEADERS:
            mapping["part_no"] = index
        elif name in DESC_HEADERS:
            mapping["description"] = index
        elif name in BOX_QTY_HEADERS or ("box" in name and "qty" in name and "pallet" not in name):
            mapping["box_qty"] = index
        elif name in PALLET_QTY_HEADERS or ("pallet" in name and "qty" in name):
            mapping["pallet_qty"] = index

    if "box_qty" not in mapping:
        for index, header in enumerate(headers):
            if header is None or str(header).strip().lower().startswith("unnamed"):
                if index not in mapping.values():
                    mapping["box_qty"] = index
                    break
    return mapping


def _cell_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        stripped = value.strip().upper()
        if not stripped or stripped == "TBA":
            return None
        if stripped.isdigit():
            return int(stripped)
        return None
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(value)
    if isinstance(value, int):
        return value
    return None


def scan_qty_for_barcode(
    barcode: str,
    manual_qty: int = 1,
) -> tuple[int, str | None]:
    """Return qty to apply for a scan and whether it came from BoxQty.

    Box barcodes apply BoxQty × ``manual_qty``. All other barcodes count as
    ``manual_qty`` singles (e.g. single-item barcodes with no BoxQty).

    PalletQty in the master list is stored for reference only — pallets are
    not scanned via barcode.

    Returns:
        (qty, qty_source) where qty_source is ``"box"`` or ``None``.
    """
    lookup = lookup_barcode(barcode)
    count = max(1, manual_qty)
    if not lookup:
        return count, None
    box_qty = lookup.get("box_qty")
    if box_qty is not None and int(box_qty) > 0:
        return int(box_qty) * count, "box"
    return count, None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank_barcode(barcode: str) -> bool:
    value = (barcode or "").strip().upper()
    return not value or value in {"TBA", "N/A", "NA", "-", "NONE"}


def _part_only_barcode(part_no: str) -> str:
    return f"__PART__{normalize_part(part_no)}"


def normalize_scanned_code(code: str) -> str:
    """Strip whitespace and common scanner prefixes before lookup."""
    value = (code or "").strip()
    if len(value) > 3 and value[0] == "]" and value[1] in "cCeE":
        value = value[3:].strip()
    return value


def load_from_excel(path: Path | None = None) -> int:
    source = path or get_master_path()
    if not source.exists():
        raise FileNotFoundError(f"Barcode master list not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = list(next(rows_iter, []))
    columns = _column_map(headers)
    if "barcode" not in columns or "part_no" not in columns:
        workbook.close()
        raise ValueError("Barcode master list must include Barcode and Item Part Number columns.")

    desc_index = columns.get("description")
    box_qty_index = columns.get("box_qty")
    pallet_qty_index = columns.get("pallet_qty")
    records: list[tuple[str, str, str, int | None, int | None]] = []
    seen: set[str] = set()
    seen_parts: set[str] = set()

    for row in rows_iter:
        if not row:
            continue
        barcode = _cell_text(row[columns["barcode"]])
        part_no = _cell_text(row[columns["part_no"]])
        if not part_no:
            continue

        part_key = normalize_part(part_no)
        if _is_blank_barcode(barcode):
            if part_key in seen_parts:
                continue
            barcode = _part_only_barcode(part_no)
        elif barcode in seen:
            continue

        if barcode in seen:
            continue
        seen.add(barcode)
        seen_parts.add(part_key)

        description = _cell_text(row[desc_index]) if desc_index is not None else ""
        box_qty = (
            _cell_int(row[box_qty_index])
            if box_qty_index is not None and box_qty_index < len(row)
            else None
        )
        if box_qty is not None and box_qty < 1:
            box_qty = None
        pallet_qty = (
            _cell_int(row[pallet_qty_index])
            if pallet_qty_index is not None and pallet_qty_index < len(row)
            else None
        )
        if pallet_qty is not None and pallet_qty < 1:
            pallet_qty = None
        records.append((barcode, part_no, description, box_qty, pallet_qty))

    workbook.close()

    with _connect() as conn:
        _ensure_tables(conn)
        conn.execute("DELETE FROM barcode_master")
        conn.executemany(
            "INSERT INTO barcode_master (barcode, part_no, description, box_qty, pallet_qty) VALUES (?, ?, ?, ?, ?)",
            records,
        )
        conn.execute(
            """
            INSERT INTO app_metadata (key, value) VALUES ('barcode_master_path', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (str(source),),
        )
        conn.execute(
            """
            INSERT INTO app_metadata (key, value) VALUES ('barcode_master_mtime', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (str(source.stat().st_mtime),),
        )
        conn.execute(
            """
            INSERT INTO app_metadata (key, value) VALUES ('barcode_master_count', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (str(len(records)),),
        )
        conn.execute(
            """
            INSERT INTO app_metadata (key, value) VALUES ('catalog_schema_version', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (CATALOG_SCHEMA_VERSION,),
        )

    _save_config_path()
    return len(records)


def import_master_file(
    source: Path | bytes,
    *,
    source_name: str = "BarcodeMasterList.xlsx",
) -> int:
    """Overwrite the app's barcode master file and reload the catalog."""
    _data_dir().mkdir(parents=True, exist_ok=True)
    master_path = _master_path()
    if isinstance(source, bytes):
        master_path.write_bytes(source)
    else:
        shutil.copy2(source, master_path)
    return load_from_excel(master_path)


def ensure_loaded() -> int:
    path = get_master_path()
    if not path.exists():
        raise FileNotFoundError(
            f"Barcode master list not found at {path}. "
            "Use 'Update Barcode List' to import BarcodeMasterList.xlsx."
        )

    file_mtime = str(path.stat().st_mtime)
    with _connect() as conn:
        _ensure_tables(conn)
        count_row = conn.execute(
            "SELECT value FROM app_metadata WHERE key = 'barcode_master_count'"
        ).fetchone()
        mtime_row = conn.execute(
            "SELECT value FROM app_metadata WHERE key = 'barcode_master_mtime'"
        ).fetchone()
        schema_row = conn.execute(
            "SELECT value FROM app_metadata WHERE key = 'catalog_schema_version'"
        ).fetchone()
        if (
            count_row
            and int(count_row["value"]) > 0
            and mtime_row
            and mtime_row["value"] == file_mtime
            and schema_row
            and schema_row["value"] == CATALOG_SCHEMA_VERSION
        ):
            return int(count_row["value"])

    return load_from_excel(path)


def catalog_count() -> int:
    with _connect() as conn:
        _ensure_tables(conn)
        row = conn.execute("SELECT COUNT(*) AS count FROM barcode_master").fetchone()
        return int(row["count"]) if row else 0


def catalog_status_text() -> str:
    path = get_master_path()
    if path.exists():
        return f"{catalog_count():,} barcodes — {path.name}"
    return "No barcode list — click Update Barcode List"


def lookup_part_no(part_no: str) -> dict[str, str] | None:
    target = normalize_part(part_no)
    if not target:
        return None

    with _connect() as conn:
        _ensure_tables(conn)
        row = conn.execute(
            """
            SELECT part_no, description
            FROM barcode_master
            WHERE UPPER(TRIM(part_no)) = ?
            LIMIT 1
            """,
            (target,),
        ).fetchone()
        if row:
            return {
                "part_no": row["part_no"],
                "description": row["description"] or "",
            }

    return None


def lookup_barcode(barcode: str) -> dict[str, str] | None:
    code = normalize_scanned_code(barcode)
    if not code:
        return None

    with _connect() as conn:
        _ensure_tables(conn)
        row = conn.execute(
            "SELECT barcode, part_no, description, box_qty, pallet_qty FROM barcode_master WHERE barcode = ?",
            (code,),
        ).fetchone()
        if row:
            result = {
                "part_no": row["part_no"],
                "description": row["description"] or "",
            }
            if row["box_qty"]:
                result["box_qty"] = int(row["box_qty"])
            if row["pallet_qty"]:
                result["pallet_qty"] = int(row["pallet_qty"])
            return result

    return None


def scanned_qty_for_part(scanned_items: list[dict[str, Any]], part_no: str) -> int:
    target = normalize_part(part_no)
    return sum(
        int(item.get("qty", 0))
        for item in scanned_items
        if normalize_part(item.get("part_no", "")) == target
        and item.get("match_status") != "unknown"
    )
